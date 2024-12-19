import openpyxl
from openpyxl.chart import PieChart, Reference


wb = openpyxl.load_workbook("excel.xlsx")
sheet = wb.active

for i in range(17, 20):
    s = sheet.cell(row=i, column=3).value.replace(" руб.", "")
    s = float(s)
    sheet.cell(row=i, column=5, value=s)

chart = PieChart()
chart.title = "Распределение зарплаты по отделам"
data_ref = Reference(sheet, min_col=5, min_row=17, max_row=19)

categories_ref = Reference(sheet, min_col=2, min_row=17, max_row=19)

chart.add_data(data_ref, titles_from_data=False)
chart.set_categories(categories_ref)

sheet.add_chart(chart, "L1")
wb.save("excel.xlsx")
